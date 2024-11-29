import frappe
from frappe.utils import flt, getdate
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

@frappe.whitelist()
def export_to_excel(filters=None):
    if not filters:
        return

    # Convertir los filtros a un diccionario si es un string JSON
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)

    # Obtener columnas y datos del reporte
    columns, data = execute(filters)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retenciones Trimestrales"

    # Estilos
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezado del reporte
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="Retenciones Trimestrales").font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Empresa
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=f"Empresa: {filters.get('company')}").font = Font(size=12, bold=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='left')

    # Tipo de Retención
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    ws.cell(row=3, column=1, value=f"Retenciones {filters.get('category')}").font = Font(size=12, bold=True)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='left')

    # Periodo
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(columns))
    ws.cell(row=4, column=1, value=f"Periodo: {filters.get('from_date')} a {filters.get('to_date')}").font = Font(size=12, bold=True)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal='left')

    # Encabezado de la tabla
    header_row = 6
    for col_num, column in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num, value=column["label"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = gray_fill
        cell.border = thin_border

    # Agregar los datos
    for row_num, row_data in enumerate(data, header_row + 1):
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data.get(column["fieldname"]))
            cell.border = thin_border

    # Guardar el archivo Excel
    file_name = f"Retenciones_Trimestrales_{filters.get('from_date')}_a_{filters.get('to_date')}.xlsx".replace(" ", "_")
    file_path = os.path.join(frappe.utils.get_site_path(), 'private', 'files', file_name)
    wb.save(file_path)

    # Guardar el archivo en la base de datos de Frappe
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "is_private": 1,
        "file_url": f"/private/files/{file_name}"
    })
    file_doc.save(ignore_permissions=True)

    return file_doc.file_url


# def execute(filters=None):
#     columns, data = get_columns_retenciones(filters), []

#     if not filters or not filters.get("category"):
#         return columns, data

#     # Obtener los filtros
#     from_date = filters.get("from_date")
#     to_date = filters.get("to_date")
#     company = filters.get("company")
#     category = filters.get("category")

#     # Diccionario con los valores de los filtros para retenciones y bases
#     filters_dict = {
#         "company": company,
#         "from_date": from_date,
#         "to_date": to_date,
#         "like_alquiler": "%alquiler%",
#         "like_profesional": "%profesional%"
#     }

#     # Condiciones SQL para la tabla GL Entry (retenciones)
#     conditions_gl = [
#         "gl.company = %(company)s",
#         "gl.posting_date >= %(from_date)s",
#         "gl.posting_date <= %(to_date)s",
#         "gl.account LIKE '4751%%'"
#     ]

#     if category == "Alquiler":
#         conditions_gl.append("LOWER(acc.account_name) LIKE %(like_alquiler)s")
#     elif category == "Profesional":
#         conditions_gl.append("LOWER(acc.account_name) LIKE %(like_profesional)s")
    
#     where_clause_gl = " AND ".join(conditions_gl)

#     # Consulta SQL para obtener las retenciones de la cuenta 4751
#     query_retenciones = f"""
#         SELECT 
#             sup.supplier_name AS proveedor,
#             sup.tax_id AS cif,
#             sup.custom_cp AS codigo_postal,
#             gl.account AS cuenta,
#             SUM(gl.credit - gl.debit) AS total_retencion
#         FROM 
#             `tabGL Entry` AS gl
#         LEFT JOIN 
#             `tabAccount` AS acc ON gl.account = acc.name
#         LEFT JOIN 
#             `tabPurchase Invoice` AS inv ON gl.voucher_no = inv.name
#         LEFT JOIN 
#             `tabSupplier` AS sup ON inv.supplier = sup.name
#         WHERE 
#             {where_clause_gl}
#         GROUP BY 
#             sup.supplier_name, sup.tax_id, sup.custom_cp, gl.account
#         ORDER BY 
#             sup.supplier_name
#     """

#     # Ejecutar la consulta para obtener las retenciones
#     retenciones = frappe.db.sql(query_retenciones, filters_dict, as_dict=True)

#     # Obtener la lista de proveedores relevantes
#     proveedores = {(retencion["proveedor"], retencion["cif"], retencion["codigo_postal"]) for retencion in retenciones}

#     # Condiciones SQL para la tabla Purchase Invoice Item (bases)
#     conditions_pii = [
#         "inv.company = %(company)s",
#         "inv.posting_date >= %(from_date)s",
#         "inv.posting_date <= %(to_date)s"
#     ]

#     if category == "Alquiler":
#         conditions_pii.append("pii.expense_account LIKE '621%%'")
#     elif category == "Profesional":
#         conditions_pii.append("pii.expense_account LIKE '623%%'")
    
#     where_clause_pii = " AND ".join(conditions_pii)

#     # Consulta SQL para obtener las bases de las cuentas 621 o 623
#     query_bases = f"""
#         SELECT 
#             sup.supplier_name AS proveedor,
#             sup.tax_id AS cif,
#             sup.custom_cp AS codigo_postal,
#             REGEXP_REPLACE(acc_base.account_name, '[0-9]', '') AS cuenta_base_unificada,
#             SUM(pii.base_net_amount) AS total_base
#         FROM 
#             `tabPurchase Invoice` AS inv
#         LEFT JOIN 
#             `tabSupplier` AS sup ON inv.supplier = sup.name
#         LEFT JOIN 
#             `tabPurchase Invoice Item` AS pii ON inv.name = pii.parent
#         LEFT JOIN 
#             `tabAccount` AS acc_base ON pii.expense_account = acc_base.name
#         WHERE 
#             {where_clause_pii}
#         GROUP BY 
#             sup.supplier_name, sup.tax_id, sup.custom_cp, cuenta_base_unificada
#         ORDER BY 
#             sup.supplier_name
#     """

#     # Ejecutar la consulta para obtener las bases
#     bases = frappe.db.sql(query_bases, filters_dict, as_dict=True)

#     # Procesar los resultados combinando retenciones y bases
#     data_dict = {}
#     for retencion in retenciones:
#         key = (retencion["proveedor"], retencion["cif"], retencion["codigo_postal"])
#         if key not in data_dict:
#             data_dict[key] = {
#                 "proveedor": retencion["proveedor"],
#                 "cif": retencion["cif"],
#                 "codigo_postal": retencion["codigo_postal"],
#                 "cuenta": retencion["cuenta"],
#                 "total": retencion["total_retencion"],
#                 "cuenta_base": None,
#                 "total_base": 0
#             }
#         else:
#             data_dict[key]["total"] += retencion["total_retencion"]

#     for base in bases:
#         key = (base["proveedor"], base["cif"], base["codigo_postal"])
#         if key in data_dict:
#             data_dict[key]["total_base"] += base["total_base"]
#             if not data_dict[key]["cuenta_base"]:
#                 data_dict[key]["cuenta_base"] = base["cuenta_base_unificada"]
#         else:
#             data_dict[key] = {
#                 "proveedor": base["proveedor"],
#                 "cif": base["cif"],
#                 "codigo_postal": base["codigo_postal"],
#                 "cuenta": None,
#                 "total": 0,
#                 "cuenta_base": base["cuenta_base_unificada"],
#                 "total_base": base["total_base"]
#             }

#     data = list(data_dict.values())

#     return columns, data

# def execute(filters=None):

#     columns, data = get_columns_retenciones(filters), []

#     if not filters or not filters.get("category"):
#         return columns, data

#     # Obtener los filtros
#     from_date = filters.get("from_date")
#     to_date = filters.get("to_date")
#     company = filters.get("company")
#     category = filters.get("category")

#     # Diccionario con los valores de los filtros para retenciones y bases
#     filters_dict = {
#         "company": company,
#         "from_date": from_date,
#         "to_date": to_date,
#         "like_alquiler": "%alquiler%",
#         "like_profesional": "%profesional%"
#     }

#     # Condiciones SQL para la tabla GL Entry (retenciones)
#     conditions_gl = [
#         "gl.company = %(company)s",
#         "gl.posting_date >= %(from_date)s",
#         "gl.posting_date <= %(to_date)s",
#         "gl.account LIKE '4751%%'"
#     ]

#     if category == "Alquiler":
#         conditions_gl.append("LOWER(acc.account_name) LIKE %(like_alquiler)s")
#     elif category == "Profesional":
#         conditions_gl.append("LOWER(acc.account_name) LIKE %(like_profesional)s")
    
#     where_clause_gl = " AND ".join(conditions_gl)

#     # Consulta SQL para obtener las retenciones de la cuenta 4751
#     query_retenciones = f"""
#         SELECT 
#             sup.supplier_name AS proveedor,
#             sup.tax_id AS cif,
#             sup.custom_cp AS codigo_postal,
#             gl.account AS cuenta,
#             SUM(gl.credit - gl.debit) AS total_retencion
#         FROM 
#             `tabGL Entry` AS gl
#         LEFT JOIN 
#             `tabAccount` AS acc ON gl.account = acc.name
#         LEFT JOIN 
#             `tabPurchase Invoice` AS inv ON gl.voucher_no = inv.name
#         LEFT JOIN 
#             `tabSupplier` AS sup ON inv.supplier = sup.name
#         WHERE 
#             {where_clause_gl}
#         GROUP BY 
#             sup.supplier_name, sup.tax_id, sup.custom_cp, gl.account
#         ORDER BY 
#             sup.supplier_name
#     """

#     # Ejecutar la consulta para obtener las retenciones
#     retenciones = frappe.db.sql(query_retenciones, filters_dict, as_dict=True)

#     # Condiciones SQL para la tabla Purchase Invoice Item (bases), usando solo las facturas que estén en GL Entry de la cuenta 4751
#     conditions_pii = [
#         "inv.company = %(company)s",
#         "inv.posting_date >= %(from_date)s",
#         "inv.posting_date <= %(to_date)s",
#         "gl.account LIKE '4751%%'",
#         "inv.docstatus = 1"  # Solo incluir facturas con estado 1 (no canceladas)
#     ]

#     if category == "Alquiler":
#         conditions_pii.append("pii.expense_account LIKE '621%%'")
#     elif category == "Profesional":
#         conditions_pii.append("pii.expense_account LIKE '623%%'")
    
#     where_clause_pii = " AND ".join(conditions_pii)

#     # Consulta SQL para obtener las bases de las cuentas 621 o 623
#     query_bases = f"""
#         SELECT 
#             sup.supplier_name AS proveedor,
#             sup.tax_id AS cif,
#             sup.custom_cp AS codigo_postal,
#             REGEXP_REPLACE(acc_base.account_name, '[0-9]', '') AS cuenta_base_unificada,
#             SUM(CASE 
#                 WHEN pii.expense_account LIKE '621%%' OR pii.expense_account LIKE '623%%'
#                 THEN pii.base_net_amount
#                 ELSE 0
#             END) AS total_base
#         FROM 
#             `tabGL Entry` AS gl
#         LEFT JOIN 
#             `tabPurchase Invoice` AS inv ON gl.voucher_no = inv.name
#         LEFT JOIN 
#             `tabSupplier` AS sup ON inv.supplier = sup.name
#         LEFT JOIN 
#             `tabPurchase Invoice Item` AS pii ON inv.name = pii.parent
#         LEFT JOIN 
#             `tabAccount` AS acc_base ON pii.expense_account = acc_base.name
#         WHERE 
#             {where_clause_pii}
#         GROUP BY 
#             sup.supplier_name, sup.tax_id, sup.custom_cp, cuenta_base_unificada
#         ORDER BY 
#             sup.supplier_name
#     """

#     # Ejecutar la consulta para obtener las bases
#     bases = frappe.db.sql(query_bases, filters_dict, as_dict=True)

#     # Procesar los resultados combinando retenciones y bases
#     data_dict = {}
#     for retencion in retenciones:
#         key = (retencion["proveedor"], retencion["cif"], retencion["codigo_postal"])
#         if key not in data_dict:
#             data_dict[key] = {
#                 "proveedor": retencion["proveedor"],
#                 "cif": retencion["cif"],
#                 "codigo_postal": retencion["codigo_postal"],
#                 "cuenta": retencion["cuenta"],
#                 "total": retencion["total_retencion"],
#                 "cuenta_base": None,
#                 "total_base": 0
#             }
#         else:
#             data_dict[key]["total"] += retencion["total_retencion"]

#     for base in bases:
#         key = (base["proveedor"], base["cif"], base["codigo_postal"])
#         if key in data_dict:
#             data_dict[key]["total_base"] += base["total_base"]
#             if not data_dict[key]["cuenta_base"]:
#                 data_dict[key]["cuenta_base"] = base["cuenta_base_unificada"]
#         else:
#             data_dict[key] = {
#                 "proveedor": base["proveedor"],
#                 "cif": base["cif"],
#                 "codigo_postal": base["codigo_postal"],
#                 "cuenta": None,
#                 "total": 0,
#                 "cuenta_base": base["cuenta_base_unificada"],
#                 "total_base": base["total_base"]
#             }

#     data = list(data_dict.values())

#     return columns, data

def execute(filters=None):
    columns, data = get_columns_retenciones(filters), []

    if not filters or not filters.get("category"):
        return columns, data

    # Obtener los filtros
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    company = filters.get("company")
    category = filters.get("category")

    # Diccionario con los valores de los filtros para retenciones y bases
    filters_dict = {
        "company": company,
        "from_date": from_date,
        "to_date": to_date,
        "like_alquiler": "%alquiler%",
        "like_profesional": "%profesional%"
    }

    # Condiciones SQL para las facturas de compra (retenciones)
    conditions_retenciones = [
        "inv.company = %(company)s",
        "inv.posting_date >= %(from_date)s",
        "inv.posting_date <= %(to_date)s",
        "tax.account_head LIKE '4751%%'",
        "inv.docstatus = 1"  # Solo incluir facturas validadas
    ]

    if category == "Alquiler":
        conditions_retenciones.append("EXISTS (SELECT 1 FROM `tabPurchase Invoice Item` AS pii WHERE pii.parent = inv.name AND pii.expense_account LIKE '621%%')")
    elif category == "Profesional":
        conditions_retenciones.append("EXISTS (SELECT 1 FROM `tabPurchase Invoice Item` AS pii WHERE pii.parent = inv.name AND pii.expense_account LIKE '623%%')")

    where_clause_retenciones = " AND ".join(conditions_retenciones)

    # Consulta SQL para obtener las retenciones asegurando que no haya duplicados en los ítems de la factura
    query_retenciones = f"""
        SELECT 
            sup.supplier_name AS proveedor,
            sup.tax_id AS cif,
            sup.custom_cp AS codigo_postal,
            tax.account_head AS cuenta,
            SUM(ABS(tax.tax_amount)) AS total_retencion  -- Usamos ABS para asegurarnos de sumar valores absolutos
        FROM 
            `tabPurchase Taxes and Charges` AS tax
        LEFT JOIN 
            `tabPurchase Invoice` AS inv ON inv.name = tax.parent
        LEFT JOIN 
            `tabSupplier` AS sup ON inv.supplier = sup.name
        WHERE 
            {where_clause_retenciones}
        GROUP BY 
            sup.supplier_name, sup.tax_id, sup.custom_cp, tax.account_head
        ORDER BY 
            sup.supplier_name
    """

    # Ejecutar la consulta para obtener las retenciones
    retenciones = frappe.db.sql(query_retenciones, filters_dict, as_dict=True)

    # Condiciones SQL para obtener las bases, asegurando que no haya duplicación de facturas
    conditions_bases = [
        "inv.company = %(company)s",
        "inv.posting_date >= %(from_date)s",
        "inv.posting_date <= %(to_date)s",
        "inv.docstatus = 1",  # Solo incluir facturas validadas
        "EXISTS (SELECT 1 FROM `tabPurchase Taxes and Charges` AS tax WHERE tax.parent = inv.name AND tax.account_head LIKE '4751%%')"  # Verificar que tenga retenciones en 4751
    ]

    if category == "Alquiler":
        conditions_bases.append("pii.expense_account LIKE '621%%'")
    elif category == "Profesional":
        conditions_bases.append("pii.expense_account LIKE '623%%'")

    where_clause_bases = " AND ".join(conditions_bases)

    # Consulta SQL para obtener las bases, unificando todas las cuentas 621 o 623
    query_bases = f"""
        SELECT 
            sup.supplier_name AS proveedor,
            sup.tax_id AS cif,
            sup.custom_cp AS codigo_postal,
            REGEXP_REPLACE(acc_base.account_name, '[0-9]', '') AS cuenta_base_unificada,  -- Unificamos el nombre de la cuenta
            SUM(ABS(pii.base_net_amount)) AS total_base
        FROM 
            `tabPurchase Invoice` AS inv
        LEFT JOIN 
            `tabPurchase Invoice Item` AS pii ON inv.name = pii.parent
        LEFT JOIN 
            `tabSupplier` AS sup ON inv.supplier = sup.name
        LEFT JOIN 
            `tabAccount` AS acc_base ON pii.expense_account = acc_base.name
        WHERE 
            {where_clause_bases}
        GROUP BY 
            sup.supplier_name, sup.tax_id, sup.custom_cp, cuenta_base_unificada
        ORDER BY 
            sup.supplier_name
    """

    # Ejecutar la consulta para obtener las bases
    bases = frappe.db.sql(query_bases, filters_dict, as_dict=True)

    # Procesar los resultados combinando retenciones y bases
    data_dict = {}
    
    # Procesamos las retenciones primero
    for retencion in retenciones:
        key = (retencion["proveedor"], retencion["cif"], retencion["codigo_postal"])
        if key not in data_dict:
            data_dict[key] = {
                "proveedor": retencion["proveedor"],
                "cif": retencion["cif"],
                "codigo_postal": retencion["codigo_postal"],
                "cuenta": retencion["cuenta"],
                "total": retencion["total_retencion"],
                "cuenta_base": None,
                "total_base": 0
            }
        else:
            data_dict[key]["total"] += retencion["total_retencion"]

    # Procesamos las bases ahora
    for base in bases:
        key = (base["proveedor"], base["cif"], base["codigo_postal"])
        if key in data_dict:
            data_dict[key]["total_base"] += base["total_base"]
            if not data_dict[key]["cuenta_base"]:
                data_dict[key]["cuenta_base"] = base["cuenta_base_unificada"]
        else:
            data_dict[key] = {
                "proveedor": base["proveedor"],
                "cif": base["cif"],
                "codigo_postal": base["codigo_postal"],
                "cuenta": None,
                "total": 0,
                "cuenta_base": base["cuenta_base_unificada"],
                "total_base": base["total_base"]
            }

    # Convertimos los resultados en una lista para devolver
    data = list(data_dict.values())

    return columns, data


def get_columns_retenciones(filters=None):
    return [
        {"label": "Proveedor", "fieldname": "proveedor", "fieldtype": "Link", "options": "Supplier", "width": 200},
        {"label": "CIF", "fieldname": "cif", "fieldtype": "Data", "width": 120},
        {"label": "Código Postal", "fieldname": "codigo_postal", "fieldtype": "Data", "width": 100},
        {"label": "Cuenta", "fieldname": "cuenta", "fieldtype": "Data", "width": 150},
        {"label": "Total Retención", "fieldname": "total", "fieldtype": "Currency", "width": 120},
        {"label": "Cuenta Base", "fieldname": "cuenta_base", "fieldtype": "Data", "width": 150},
        {"label": "Total Base", "fieldname": "total_base", "fieldtype": "Currency", "width": 120}
    ]