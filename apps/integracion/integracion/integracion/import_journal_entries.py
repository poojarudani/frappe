import frappe

import openpyxl
from openpyxl import load_workbook

import os
from datetime import datetime
import logging

# Configurar el logger
logger = logging.getLogger(__name__)

handler = logging.FileHandler(
    '/home/frappe/frappe-bench/apps/integracion/integracion/integracion/logs/import_journal_entries.log'
)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)

@frappe.whitelist()
def get_progress():
	progreso = frappe.cache().hget('import_journal_entries_progreso', 'progreso')
	total_asientos = frappe.cache().hget('import_journal_entries_progreso', 'total_asientos')
	estado = frappe.cache().hget('import_journal_entries_progreso', 'estado')

	if progreso and total_asientos and estado:
		return {
            "progreso": int(progreso),
            "total_asientos": int(total_asientos),
			"estado": estado
        }

	return None

@frappe.whitelist()
def import_journal_entries(excel_file, company: str):
	# Cargar progreso al cache Frappe

	progreso = frappe.cache().hset('import_journal_entries_progreso', 'progreso', 0)
	total_asientos = frappe.cache().hset('import_journal_entries_progreso', 'total_asientos', 100)
	estado = frappe.cache().hset('import_journal_entries_progreso', 'estado', "Limpiando datos...")
	company_abbr = frappe.db.get_value("Company", company, "abbr")

	# Cargar archivo subido
	file_doc = frappe.get_doc("File", {"file_url": excel_file})
	wb = load_workbook(file_doc.get_full_path(), data_only=True)

	# Datos limpios
	data = []

	# Datos errados
	errors_header = []
	errors_rows = []

	# Transferir cabecera del archivo original al Excel de errores de importación (Es siempre hasta la fila 7)
	for row in list(wb.active.iter_rows())[:7]:
		errors_header.append([c.value for c in row])

	# Limpieza de datos y recolección de errores
	for ws in wb.worksheets:

		# Todas las filas después de la cabecera (Donde empiezan los datos)
		ws_rows = list(ws.iter_rows())[7:]

		# Filtrar todas las filas con numero de apunte
		ws_rows = list(filter(lambda r: r[2].value != None, ws_rows))

		# ACTUALIZAR PROGRESO
		## Cambiar estado
		estado = f"Limpiando datos hoja {ws.title}"
		frappe.cache().hset('import_journal_entries_progreso', 'total_asientos', len(ws_rows))
		frappe.cache().hset('import_journal_entries_progreso', 'progreso', 1)

		## Sumar cantidad de filas de la hoja al total de asientos
		total_asientos = frappe.cache().hget('import_journal_entries_progreso', 'total_asientos')

		date = None
		next_journal_entry = False
		numero_asiento = None
		journal_entry = {"lines": [], "errors": False}

		total_debit = 0
		total_credit = 0

		logger.info(f"HOJA: {ws}")
		logger.info("*"*70)

		# Recorrer cada fila de la hoja
		for row in ws_rows:
			cells = [cell.value for cell in row[:9]]

			# Agregar Party type y Party como celdas extra, por defecto None
			cells.extend([None, None])

			# Mapeo de celdas
			numero_asiento = cells[1]
			numero_apunte = cells[2]
			cuenta = cells[5]
			debit = cells[7]
			credit = cells[8]

			logger.info(f"Asiento: {numero_asiento} Apunte: {numero_apunte}")

			# Si la fila es cabecera y es momento de cambiar de cabecera
			if numero_asiento and next_journal_entry:

				# Verificar si el debe y haber del asiento coinciden
				if round(total_credit, 2) != round(total_debit, 2):
					journal_entry["lines"][0].append("Total debe y total haber no coinciden.")
					journal_entry["errors"] = True

				# Si el asiento tienes errores, agregar a las filas erróneas
				if journal_entry["errors"]:
					errors_rows.append(journal_entry["lines"])
				else:
					# Si no, agregar a los datos limpios
					data.extend(journal_entry["lines"])

				# Reiniciar estado de error y de la cabecera del asiento
				journal_entry["errors"] = False
				next_journal_entry = False

			# Verificar si la celda tiene un formato de fecha válido
			if cells[0] != None:
				# Si es datetime, cambiar el valor de la fecha
				if type(cells[0]) == datetime:
					date = cells[0]
				else:
					# Si no, intentar convertirla
					try:
						cells[0] = datetime.strptime(cells[0], "%d-%b.-%y")
						date = cells[0]
					except Exception as e:					
						# Agregar error al asiento
						cells.append(f"Error al formatear fecha | {e}")
						journal_entry["lines"].append(cells)
						journal_entry["errors"] = True

			# Buscar DocType de la cuenta con el número de cuenta del excel original
			account_doc = frappe.db.get_value("Account", {"account_number": cuenta, "company": company}, "name")

			# Si lo encuentra, cambiar el valor de la celda de cuenta
			if account_doc:
				cells[5] = account_doc
			else:
				# Si no, buscar cuenta por el nombre de la cuenta
				account_doc = frappe.db.get_value("Account", cuenta, "name")

				# Si lo encuentra, cambiar el valor de la celda de cuenta
				if account_doc:
					cells[5] = account_doc
				else:
					# Si no, agregar error al asiento
					cells.append(
						f"""Cuenta "{cuenta}" no existe en compañía "{company}". Verificar el plan contable."""
					)
					journal_entry["errors"] = True

			# Buscar Party Type y Party de la cuenta
			if account_doc:
				logger.info(f"Buscando Party Type y Party para {account_doc}...")

				account_name, cuenta_numero, padre, account_type = frappe.db.get_value(
					"Account", account_doc, ["account_name", "account_number", "parent_account", "account_type"]
				)
				padre_numero = None

				logger.info(f"TIPO DE CUENTA {account_type}")

				if padre:
					padre_numero = frappe.db.get_value("Account", padre, "account_number")
					logger.info(f"Cuenta padre {padre}")

				if cuenta_numero == "40000000" or cuenta_numero == "40700200" or cuenta_numero == "40701000":
					logger.info("""Buscando proveedor "movimientos importados"...""")

					party_type = "Supplier"

					supplier = frappe.db.get_value(party_type, {"name": "Movimientos Importados"})

					if supplier:
						cells[9] = party_type
						cells[10] = supplier
					else:
						cells.append(
							f"""No se ha encontrado el proveedor "Movimientos Importados" en compañía {company}."""
						)
						journal_entry["errors"] = True

				if cuenta_numero == "46500000" or cuenta_numero == "46500010" or cuenta_numero == "465000000":
					logger.info("""Buscando empleado "movimientos importados"...""")

					employee = frappe.db.get_value(
						"Employee", {"employee_name": ("like", "%Movimientos Importados%")}
					)

					if employee:
						cells[9] = "Employee"
						cells[10] = employee
					else:
						cells.append(
							f"""No se ha encontrado el empleado "Movimientos Importados" en compañía {company}."""
						)
						journal_entry["errors"] = True

				if padre_numero == "4100" or padre_numero == "410" or padre_numero == "4000":
					party_type = "Supplier"

					if cuenta_numero == "41000000" or cuenta_numero == "410000000":
						supplier = frappe.db.get_value(party_type, {"name": "Movimientos Importados"})

						if supplier:
							cells[9] = party_type
							cells[10] = supplier
						else:
							cells.append(
								f"""No se ha encontrado el proveedor "Movimientos Importados" en compañía {company}."""
							)
							journal_entry["errors"] = True

					else:
						supplier_name = account_name

						if account_name.find(" - ") != -1:
							supplier_name = account_name.split(" - ")[1]
							supplier_name = account_name.rstrip()

						supplier = frappe.db.get_value(
							party_type, {"name": ("like", f"%{supplier_name}%")}
						)

						if not supplier:
							supplier = frappe.db.get_value("Party Account", {"account": account_name}, "parent")

							if not supplier:
								supplier = frappe.db.get_value(party_type, {"name": "Movimientos Importados"})

						if supplier:
							cells[9] = party_type
							cells[10] = supplier

				if padre_numero == "4300" or padre_numero == "430" or padre_numero == "440":
					logger.info(f"""Buscando cliente para cuenta "{account_name}"...""")
					party_type = "Customer"

					customer_name = account_name

					if customer_name.find(" - ") != -1:
						logger.info("Limpiando guiones")
						customer_name = customer_name.split(" - ")[1]
						customer_name = customer_name.rstrip()
						logger.info(f"Resultado: {customer_name}")

					customer = frappe.db.get_value(
						party_type, {"customer_name": ("like", f"%{customer_name}%")}
					)

					if not customer:
						customer = frappe.db.get_value("Party Account", {"account": account_name}, "parent")

						if not customer:
							customer = frappe.db.get_value(party_type, {"name": "Movimientos Importados"})

					if customer:
						cells[9] = party_type
						cells[10] = customer

				if account_type in ("Payable", "Receivable") and (not cells[9] and not cells[10]):
					party_type = None
					party = None

					party_account = frappe.db.get_value(
						"Party Account", {"account": account_name}, ["parenttype", "parent"]
					)

					if party_account:
						party_type = party_account[0]
						party = party_account[1]

					else:
						if account_type == "Receivable":
							party_type = "Customer"
							customer_name = account_doc

							if customer_name.find(" - ") != -1:
								logger.info("Limpiando guiones")
								customer_name = customer_name.split(" - ")[1]
								customer_name = customer_name.rstrip()
								logger.info(f"Resultado: {customer_name}")

							party = frappe.db.get_value(party_type, {"customer_name": ("like", f"%{customer_name}%")})

						elif account_type == "Payable":
							party_type = "Supplier"
							supplier_name = account_doc

							if supplier_name.find(" - ") != -1:
								logger.info("Limpiando guiones")
								supplier_name = supplier_name.split(" - ")[1]
								supplier_name = supplier_name.rstrip()
								logger.info(f"Resultado: {supplier_name}")

							party = frappe.db.get_value(
								party_type, {"supplier_name": ("like", f"%{supplier_name}%")}
							)

					if party_type and party:
						cells[9] = party_type
						cells[10] = party
					else:
						cells.append(
							f"""No se ha encontrado un proveedor/cliente para la cuenta "{account_doc}"."""
						)

						journal_entry["errors"] = True
			if date:

				# Verificar si es cabecera
				if numero_asiento:

					# Reiniciar valores de total crédito y total débito
					total_credit = credit
					total_debit = debit

					journal_entry["lines"] = [cells]
				else:

					# Si no, es línea
					next_journal_entry = True

					# Suma total de crédito y débito
					total_credit += credit
					total_debit += debit

					# Agregar celdas a las líneas de asientos
					journal_entry["lines"].append(cells)

					# Si la fila está ubicada al final de la hoja, verificar si las líneas no tienen errores
					if row == ws_rows[-1]:
						if round(total_credit, 2) != round(total_debit, 2):
							journal_entry["lines"][0].append("Total debe y total haber no coinciden.")
							journal_entry["errors"] = True

						if journal_entry["errors"]:
							errors_rows.append(journal_entry["lines"])
						else:
							data.extend(journal_entry["lines"])

						journal_entry["errors"] = False

			else:

				# Si no tiene campo de fecha, agregar error al asiento
				if numero_asiento:
					cells.append("Fecha no especificada.")
					journal_entry["errors"] = True 

			logger.info(f"""Tiene errores {journal_entry["errors"]}""")
			logger.info("_"*70)

			# Actualizar progreso
			progreso = frappe.cache().hget('import_journal_entries_progreso', 'progreso')
			progreso += 1
			frappe.cache().hset('import_journal_entries_progreso', 'progreso', progreso)

			frappe.publish_realtime(
				"import_journal_entries_progreso", {
					"progress": [progreso, total_asientos], "message": estado, "success": False
				},
				user=frappe.session.user
			)

	# Booleano para verificar si está listo el asiento para guardarlo en Frappe
	commit_journal_entry = False

	# Variable para almacenar cabecera y líneas del asiento a guardar en Frappe
	journal_entry = None

	# Recorrer datos limpios para subirlos a asientos aontables en Frappe

	progreso = 0
	total_asientos = len(data)

	for cells in data:
		progreso += 1

		logger.info(f"Creando... {cells}")

		frappe.publish_realtime(
			"import_journal_entries_progreso", {
				"progress": [progreso, total_asientos], 
				"message": f"Creando línea de asiento {progreso} de {total_asientos}",
				"success": False
			},
			user=frappe.session.user
		)

		# Mapeo campos de fila
		date = cells[0]
		numero_asiento = cells[1]
		concepto = cells[3]
		documento = cells[4]
		cuenta = cells[5]
		debit = cells[7]
		credit = cells[8]
		party_type = cells[9]
		party = cells[10]

		# Verificar si es apertura
		apertura = False

		if documento and documento.lower() == "apertura":
			apertura = True

		# Si tiene número asiento es cabecera
		if numero_asiento != None:

			# Chequear si tiene que guardar Journal Entry anterior
			if commit_journal_entry:
				journal_entry.insert(ignore_permissions=True)
				frappe.db.commit()
				commit_journal_entry = False

			# Crear cabecera de asiento con primera línea
			journal_entry = frappe.get_doc({
				"doctype": "Journal Entry",

				# Cambiar a Opening Entry si es apertura
				"voucher_type": "Opening Entry" if apertura else "Journal Entry",
				"title": f"""Apertura {numero_asiento} - {date.strftime("%d/%m/%Y")} {concepto} - {company_abbr}"""[:139] if apertura else f"""Asiento {numero_asiento} - {date.strftime("%d/%m/%Y")} {concepto} - {company_abbr}"""[:139],
				"company": company,
				"number": numero_asiento,
				"posting_date": date,
				"is_opening": apertura,
				"accounts": [{
					"account": cuenta,
					"party_type": party_type,
					"party": party,
					"debit_in_account_currency": debit,
					"credit_in_account_currency": credit,
				}]
			})

		else:
			# Agregar línea a cabecera anteriormente creada
			journal_entry.append("accounts", {
				"account": cuenta,
				"party_type": party_type,
				"party": party,
				"debit_in_account_currency": debit,
				"credit_in_account_currency": credit,
			})

			# Si la línea se encuentra al final de la hoja, guardar asiento actual a asientos contables de Frappe
			if cells == data[-1]:
				journal_entry.insert(ignore_permissions=True)
				frappe.db.commit()
			else:
				commit_journal_entry = True

	# Borrar cache cuando termine el proceso
	frappe.cache().hdel('import_journal_entries_progreso', 'progreso')
	frappe.cache().hdel('import_journal_entries_progreso', 'total_asientos')
	frappe.cache().hdel('import_journal_entries_progreso', 'estado')

	# Si encontró errores retornar el excel de errores
	res = {"errores": len(errors_rows), "asientos": len(data), "success": True, "error_file": None}

	if errors_rows:
		frappe.publish_realtime(
			"import_journal_entries_progreso", {
				"progress": [2, 4], "message": "Generando Excel de errores", "success": False
			},
			user=frappe.session.user
		)

		errores_excel = gen_errors_excel(errors_header, errors_rows)

		res["error_file"] = errores_excel

	frappe.publish_realtime(
		"import_journal_entries_progreso",
		res,
		user=frappe.session.user
	)

	return res["error_file"]

def gen_errors_excel(header: list, data: list) -> str:
	# Crear un archivo Excel
	wb = openpyxl.Workbook()
	ws = wb.active
	title = "correccion_cuentas_importar"
	ws.title = title

	# Agregar campos de Party en la cabecera del Excel
	if header[0][1] != title:
		header[0][1] = title
		header[6].extend(["Tipo de entidad", "Tercero", "Corregir"])

	row = 0

	# Escribir cabecera
	for header_row in header:
		row += 1
		column = 0

		for cell in header_row:
			column += 1
			ws.cell(row=row, column=column, value=cell)

	# Escribir líneas
	for row_data in data:
		for line in row_data:
			row += 1
			column = 0

			for cell in line:
				column += 1
				ws.cell(row=row, column=column, value=cell)

	# Generar nombre del archivo Excel
	file_name = "correcciones_importacion_cuentas.xlsx"
	file_path = os.path.join(frappe.utils.get_site_path(), 'private', 'files', file_name)
	wb.save(file_path)

	# Guardar el archivo en la base de datos de Frappe
	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": file_name,
		"is_private": 1,
		"file_url": f"/private/files/{file_name}"
	})

	file_doc.save(ignore_permissions=True)

	return file_doc.file_url