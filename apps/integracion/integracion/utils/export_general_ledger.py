import os
import frappe
from frappe.utils.pdf import get_pdf
from frappe.utils.jinja import render_template
from frappe import _
import json
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging

# Configurar el logger
logger = logging.getLogger(__name__)

handler = logging.FileHandler(
    '/home/frappe/frappe-bench/apps/integracion/integracion/integracion/logs/export_general_ledger.log'
)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)


@frappe.whitelist()
def export_general_ledger(format, filters):
    if isinstance(filters, str):
        filters = json.loads(filters)

    # Obtener los datos del General Ledger según los filtros
    general_ledger_data = get_general_ledger_data(filters)


    # Definir el nombre basado en la cuenta, party o compañía
    account_name = filters.get("account")[0] if filters.get("account") else None
    company_name = filters.get("company", "")
    party_name = filters.get("party_name", "")

    # Si no hay cuenta y hay party_name, buscar la cuenta predeterminada
    if not account_name and party_name:
        account_name = get_default_account_for_party(party_name, company_name)

    # Si no se encontró cuenta, usar el party_name o el nombre de la compañía
    document_name = account_name or party_name or company_name


    # Obtener fecha de hoy
    today_date = datetime.date.today().strftime("%d/%m/%Y")

    # Formatear periodo
    from_date = datetime.datetime.strptime(filters.get("from_date"), "%Y-%m-%d").strftime("%d-%b")
    to_date = datetime.datetime.strptime(filters.get("to_date"), "%Y-%m-%d").strftime("%d-%b del %Y")
    formatted_today = datetime.datetime.strptime(today_date, "%d/%m/%Y").strftime("%d-%b del %Y")
    formatted_period = f"De {from_date} a {to_date}"

    # Totales
    total_debit = round(sum([entry.get('debit', 0) for entry in general_ledger_data]), 3)
    total_credit = round(sum([entry.get('credit', 0) for entry in general_ledger_data]), 3)
    total_balance = round(total_debit - total_credit, 3)

    if format == "PDF":
        # Contenido del encabezado directamente en el .py
        header_html = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: 'Arial', sans-serif;
                    font-size: 10px;
                    margin: 0;
                    padding: 0;
                    position: relative;
                }}
                .header {{
                    margin-bottom: 5px;
                }}
                .header h1 {{
                    font-size: 16px;
                    text-align: left;
                    margin: 0;
                    padding-bottom: 5px;
                }}
                .company-info {{
                    display: flex;
                    justify-content: space-between;
                    font-size: 10px;
                    font-weight: bold;
                }}
                .company-info .left {{
                    padding-left: 10px;
                    text-align: left;
                }}
                .company-info .right {{
                    text-align: right;
                    padding-right: 10px;
                }}
                .page-info {{
                    display: flex;
                    justify-content: space-between;
                    font-size: 10px;
                    font-weight: bold;
                    margin-top: 5px;
                }}
                .page-info .left {{
                    text-align: left;
                    padding-left: 10px;
                }}
                .page-info .right {{
                    text-align: right;
                    padding-right: 10px;
                }}
                .divider {{
                    border-top: 1px solid black;
                    margin: 5px 0;
                }}
                .observations {{
                    font-size: 10px;
                    text-align: center;
                    margin: 5px 10px;
                    padding: 5px;
                    background-color: #f2f2f2;
                    border: 1px solid black;
                }}
                .table-header {{
                    border-bottom: 1px solid #000;
                    font-weight: bold;
                    display: flex;
                    justify-content: space-between;
                    font-size: 10px;
                    padding: 3px 0;
                }}
                .table-header span {{
                    display: inline-block;
                    text-align: center;
                    width: 12%;  /* Ajustado para mejorar el espaciado */
                }}
                .footer {{
                    position: absolute;
                    bottom: 20px;
                    left: 0;
                    right: 0;
                    text-align: center;
                    font-size: 8px;
                    margin-top: 10px;
                    padding-top: 5px;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Listado de Cuentas Corrientes</h1>
                <div class="divider"></div>
                <div class="company-info">
                    <div class="left">Empresa: {company_name}</div>
                </div>
                <div class="page-info">
                    <div class="left">Observaciones</div>
                    <div class="right">Fecha listado: {formatted_today}</div>
                </div>
                <div class="page-info">
                    <div class="right">Periodo: {formatted_period}</div>
                </div>
                <div class="divider"></div>
                <div class="observations">
                    {document_name}
                </div>
                <div class="table-header">
                    <span>Fecha Reg.</span>
                    <span>Fecha Fac.</span>
                    <span>Concepto</span>
                    <span>Documento</span>
                    <span>Importe Debe</span>
                    <span>Importe Haber</span>
                    <span>Saldo Contrapar.</span>
                    <span>Contra cuenta</span>
                </div>
            </div>
        </body>
        </html>
        """


        # Contenido del cuerpo directamente en el .py
        body_html = """
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Libro Mayor</title>
            <style>
                body {
                    font-family: 'Arial', sans-serif;
                    font-size: 10px;
                }
                .entry {
                    margin-bottom: 2px;
                    display: flex;
                    justify-content: space-between;
                }
                .entry span {
                    display: inline-block;
                    text-align: center;
                    width: 12%;
                }
                .totals {
                    margin-top: 10px;
                    border-top: 1px solid #000;
                    font-weight: bold;
                    display: flex;
                    justify-content: space.between;
                }
                .totals .total {
                    margin-top: 10px;
                    display: inline-block;
                    text-align:center;
                    width: 12%;
                }
            </style>
        </head>
        <body>
        """

        # Añadir las entradas de la tabla
        debit_style = ''
        credit_style = ''
        balance_style = ''

        # Saldo acumulado
        acum_balance = 0

        # Añadir las entradas de la tabla
        for entry in general_ledger_data:
            
            # Estilos para valores negativos
            debit_style = 'color: red;' if entry['debit'] < 0 else ''
            credit_style = 'color: red;' if entry['credit'] < 0 else ''
            balance_style = 'color: red;' if entry['balance'] < 0 else ''

            body_html += f"""
            <div class="entry">
                <span>{entry['posting_date']}</span>
                <span>{entry['bill_date'] if entry['bill_date'] else ""}</span>
                <span>{entry['concept']}</span>
                <span>{entry['bill_no']}</span>
                <span style="{debit_style}">{entry['debit']}</span>
                <span style="{credit_style}">{entry['credit']}</span>
                <span style="{balance_style}">{round(entry['balance'] + acum_balance, 3)}</span>
                <span>{entry["against"]}</span>
            </div>
            """

            acum_balance += round(entry["balance"], 3)

        # Añadir los totales al final del cuerpo
        body_html += f"""
        <div class="totals">
            <span class="total"></span>
            <span class="total"></span>
            <span class="total"></span>
            <span class="total">Total cuenta</span>
            <span class="total" style="{debit_style}">{total_debit}</span>
            <span class="total" style="{credit_style}">{total_credit}</span>
            <span class="total" style="{balance_style}">{total_balance}</span>
        </div>
        </body>
        </html>
        """

        # Combinar el contenido del encabezado y el cuerpo
        html_content = header_html + body_html

        # Generar el archivo PDF
        pdf_content = get_pdf(html_content, {
            "orientation": "Landscape",
            #"header-spacing": 5,
            "footer-spacing": 5,
            "footer-right": "Página: [page] de [toPage]",
            "footer-font-size": 10,
        })


        # Generar nombre del archivo PDF
        document_name = document_name.replace(".", "_")
        document_name = document_name.replace(" ", "_")
        document_name = document_name.replace("%", "_")

        file_name = f"{document_name}_Libro_Mayor.pdf"
        file_path = os.path.join(frappe.utils.get_site_path(), 'private', 'files', file_name)

        # Guardar el archivo PDF
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "is_private": 1,
            "content": pdf_content,
        })
        file_doc.save(ignore_permissions=True)

        return file_doc.file_url


    if format == "Excel":
        # Crear un archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Mayor General"

        # Estilos
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Gris claro
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Margen superior e izquierdo de 2 filas y 2 columnas
        start_row = 3
        start_col = 3  # Esto corresponde a la columna "C"

        def apply_border_to_range(ws, start_row, start_col, end_row, end_col, border):
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).border = border


        def apply_alignment_to_range(ws, start_row, start_col, end_row, end_col, alignment):
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).alignment = alignment

        # Configurar encabezado simulado
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 7)
        ws.cell(row=start_row, column=start_col, value="Listado de Cuentas Corrientes")
        ws[f"C{start_row}"].font = Font(size=14, bold=True)
        ws[f"C{start_row}"].alignment = Alignment(horizontal='center')
        ws[f"C{start_row}"].fill = gray_fill
        apply_border_to_range(ws, start_row, start_col, start_row, start_col + 7, thin_border)

        ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=start_col + 7)
        ws.cell(row=start_row + 1, column=start_col, value=f"Empresa: {filters.get('company')}")
        ws[f"C{start_row + 1}"].font = Font(size=12, bold=True)
        ws[f"C{start_row + 1}"].alignment = Alignment(horizontal='left')

        ws.merge_cells(start_row=start_row + 2, start_column=start_col, end_row=start_row + 2, end_column=start_col + 3)
        ws.cell(row=start_row + 2, column=start_col, value=f"Fecha listado: {formatted_today}")
        apply_alignment_to_range(ws, start_row + 2, start_col, start_row + 2, start_col + 3, Alignment(horizontal='left'))

        ws.merge_cells(start_row=start_row + 2, start_column=start_col + 4, end_row=start_row + 2, end_column=start_col + 7)
        ws.cell(row=start_row + 2, column=start_col + 4, value=f"Periodo: {formatted_period}")
        apply_alignment_to_range(ws, start_row + 2, start_col + 4, start_row + 2, start_col + 7, Alignment(horizontal='right'))

        ws.merge_cells(start_row=start_row + 3, start_column=start_col, end_row=start_row + 3, end_column=start_col + 7)
        ws.cell(row=start_row + 3, column=start_col, value=f"Proveedor: {filters.get('party_name')}")
        ws[f"C{start_row + 3}"].font = Font(size=12, bold=True)
        ws[f"C{start_row + 3}"].alignment = Alignment(horizontal='right')

        # Celda con el account o party
        ws.merge_cells(start_row=start_row + 4, start_column=start_col, end_row=start_row + 4, end_column=start_col + 7)
        account_info = f"{filters.get('account')[0] if filters.get('account') else filters.get('party_name', filters.get('company'))}"
        ws.cell(row=start_row + 4, column=start_col, value=account_info)
        ws[f"C{start_row + 4}"].font = Font(size=14, bold=True)
        ws[f"C{start_row + 4}"].alignment = Alignment(horizontal='center')
        ws[f"C{start_row + 4}"].fill = gray_fill
        apply_border_to_range(ws, start_row + 4, start_col, start_row + 4, start_col + 7, thin_border)

        # Encabezado de la tabla (ajustamos fila)
        headers = ["Fecha Reg.", "Fecha Fac.", "Concepto", "Documento", "Debe", "Haber", "Saldo", "Contra cuenta"]
        header_row = start_row + 6
        for col_num, header in enumerate(headers, start_col):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{header_row}"] = header
            ws[f"{col_letter}{header_row}"].font = Font(bold=True)
            ws[f"{col_letter}{header_row}"].alignment = Alignment(horizontal='center')

        # Aplicar formato condicional para valores negativos
        negative_font = Font(color="FF0000")  # Rojo para los valores negativos

        # Agregar los datos del General Ledger (empezando en la fila 10)
        row_num = header_row + 1

        # Saldo acumulado
        acum_balance = 0

        for entry in general_ledger_data:
            ws[f"C{row_num}"] = entry['posting_date']
            ws[f"D{row_num}"] = entry['bill_date'] if entry['bill_date'] else ""
            ws[f"E{row_num}"] = entry['concept']
            
            # Columna Documento con condición de Purchase Invoice
            if entry['concept'] == _("Purchase Invoice") and entry.get("bill_no"):
                ws[f"F{row_num}"] = entry["bill_no"]
            else:
                ws[f"F{row_num}"] = entry['voucher_no']

            # Columna Debit
            ws[f"G{row_num}"] = entry['debit']
            if entry['debit'] < 0:
                ws[f"G{row_num}"].font = negative_font  # Aplicar el estilo si es negativo

            # Columna Credit
            ws[f"H{row_num}"] = entry['credit']
            if entry['credit'] < 0:
                ws[f"H{row_num}"].font = negative_font  # Aplicar el estilo si es negativo

            # Columna Balance
            ws[f"I{row_num}"] = round(entry['balance'] + acum_balance, 3)
            if entry['balance'] < 0:
                ws[f"I{row_num}"].font = negative_font  # Aplicar el estilo si es negativo

            ws[f"J{row_num}"] = entry["against"]

            acum_balance += round(entry['balance'], 3)
            row_num += 1

        # Agregar Totales
        ws[f"C{row_num}"] = "Total"
        ws[f"C{row_num}"].font = Font(bold=True)

        ws[f"G{row_num}"] = total_debit
        if ws[f"G{row_num}"].value < 0:
            ws[f"G{row_num}"].font = negative_font

        ws[f"H{row_num}"] = total_credit
        if ws[f"H{row_num}"].value < 0:
            ws[f"H{row_num}"].font = negative_font

        ws[f"I{row_num}"] = total_balance
        if ws[f"I{row_num}"].value < 0:
            ws[f"I{row_num}"].font = negative_font

        # Ajustar ancho de columnas
        for col in range(start_col, start_col + 7):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Generar nombre del archivo Excel
        document_name = f"Libro_Mayor_{filters.get('account')[0] if filters.get('account') else filters.get('party_name', filters.get('company'))}"
        document_name = document_name.replace(".", "_")
        document_name = document_name.replace(" ", "_")
        document_name = document_name.replace("%", "_")

        file_name = f"{document_name}.xlsx"
        file_path = os.path.join(frappe.utils.get_site_path(), 'private', 'files', file_name)
        wb.save(file_path)

        # Guardar el archivo en la base de datos de Frappe
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "is_private": 1,
            "file_url": f"/private/files/{file_name}"
        })
        file_doc.save(ignore_permissions=True)

        return file_doc.file_url

# Función auxiliar para obtener los datos del General Ledger
def get_general_ledger_data(filters):
    query_conditions = []
    query_filters = {
        "company": filters.get("company"),
        "from_date": filters.get("from_date"),
        "to_date": filters.get("to_date"),
    }
    
    query_conditions.append("company = %(company)s")
    query_conditions.append("posting_date BETWEEN %(from_date)s AND %(to_date)s")

    if filters.get("party"):
        query_conditions.append("party IN %(party)s")
        query_filters["party"] = tuple(filters.get("party"))
    
    if filters.get("account"):
        query_conditions.append("account IN %(account)s")
        query_filters["account"] = tuple(filters.get("account"))
    
    query = f"""
        SELECT posting_date, account, voucher_no, party_type, party, debit, credit,
               (debit - credit) AS balance, against
        FROM `tabGL Entry`
        WHERE {" AND ".join(query_conditions)}
        ORDER BY posting_date ASC
    """
    
    # Ejecutar la consulta
    raw_data = frappe.db.sql(query, query_filters, as_dict=True)
    
    # Crear una nueva lista para almacenar solo las entradas válidas
    filtered_data = []

    for entry in raw_data:
        entry["bill_date"] = None
        entry["bill_no"] = entry["voucher_no"]  # Default to voucher_no if not Purchase Invoice

        try:
            # Obtener el tipo de documento del voucher filtrando por voucher_no y account
            voucher_doctype = frappe.db.get_value("GL Entry", {"voucher_no": entry["voucher_no"], "account": entry["account"]}, "voucher_type")

            if voucher_doctype:
                # Verificar docstatus del documento en el Doctype específico
                docstatus = frappe.db.get_value(voucher_doctype, entry["voucher_no"], "docstatus")
                
                # Solo incluir si docstatus es 1 (enviado/aprobado)
                if docstatus == 1:
                    # Traducir el nombre del Doctype
                    translated_doctype = _(voucher_doctype)
                    entry["concept"] = translated_doctype

                    # Si es factura de compra, obtener bill_date y bill_no
                    if voucher_doctype == "Purchase Invoice":
                        entry["bill_date"] = frappe.db.get_value("Purchase Invoice", entry["voucher_no"], "bill_date")
                        entry["bill_no"] = frappe.db.get_value("Purchase Invoice", entry["voucher_no"], "bill_no") or entry["voucher_no"]

                    # Agregar entrada a la lista final si pasa todas las condiciones
                    filtered_data.append(entry)
                # Si el docstatus no es 1, se ignora la entrada (no se agrega a filtered_data)
            else:
                entry["concept"] = _("Unknown")
                entry["bill_no"] = entry["voucher_no"]
                # Si no se encuentra un Doctype válido, también se ignora la entrada
        except Exception as e:
            # Loguear cualquier excepción que ocurra en este bloque
            frappe.log_error(f"Error processing voucher_no {entry['voucher_no']}: {str(e)}", "Voucher Processing Error")
            entry["concept"] = _("Unknown")

    # Filtro por Fecha Factura
    if filters.get("bill_date"):
        filtered_data = list(filter(
            lambda e: e["bill_date"] == datetime.datetime.strptime(filters["bill_date"], "%Y-%m-%d").date(),
            filtered_data
        ))

    return filtered_data


def get_default_account_for_party(party_name, company_name):
    """
    Obtener la cuenta predeterminada para un party (Cliente o Proveedor) en base al nombre del party y la compañía.
    """
    default_account = None
    try:
        # Buscar el registro de Party Account que coincide con el party y la compañía
        default_account = frappe.db.get_value("Party Account", {
            "parent": party_name,
            "company": company_name
        }, "account")
    except Exception as e:
        frappe.log_error(f"Error fetching default account for party {party_name}: {str(e)}", "Default Account Fetch Error")
    
    return default_account