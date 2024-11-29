# Copyright (c) 2024, Xappiens and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import flt, getdate
import re
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@frappe.whitelist()
def export_to_excel(filters=None):
    """
    Exporta el informe generado a un archivo Excel.
    
    Args:
        filters (dict | None): Filtros utilizados para generar el informe.
    
    Returns:
        str: URL del archivo Excel exportado.
    """
    if not filters:
        return

    # Convertir los filtros a un diccionario si es un string JSON
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)

    # Obtener columnas y datos del reporte
    columns, data = execute(filters)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de IVA"

    # Estilos
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezado del reporte
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="Reporte de IVA").font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Empresa
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=f"Empresa: {filters.get('company')}").font = Font(size=12, bold=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='left')

    # Periodo
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    ws.cell(row=3, column=1, value=f"Periodo: {filters.get('from_date')} a {filters.get('to_date')}").font = Font(size=12, bold=True)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='left')

    # Tipo de Factura
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(columns))
    ws.cell(row=4, column=1, value=f"Tipo de Factura: {'Factura de Compra' if filters.get('type') == 'Factura de Compra' else 'Factura de Venta'}").font = Font(size=12, bold=True)
    ws.cell(row=4, column=1).alignment = Alignment(horizontal='left')

    # Porcentaje de IVA seleccionado
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(columns))
    ws.cell(row=5, column=1, value=f"Porcentaje de IVA seleccionado: {filters.get('iva_type')}%").font = Font(size=12, bold=True)
    ws.cell(row=5, column=1).alignment = Alignment(horizontal='left')

    # Valor (Positiva o Negativa)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(columns))
    ws.cell(row=6, column=1, value=f"Valor: {filters.get('valor')}").font = Font(size=12, bold=True)
    ws.cell(row=6, column=1).alignment = Alignment(horizontal='left')

    # Encabezado de la tabla
    header_row = 8
    for col_num, column in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num, value=column["label"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = gray_fill
        cell.border = thin_border

    # Agregar los datos
    for row_num, row_data in enumerate(data, header_row + 1):
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data.get(column["fieldname"]))
            cell.border = thin_border

    # Guardar el archivo Excel
    file_name = f"Reporte_IVA_{filters.get('from_date')}_a_{filters.get('to_date')}.xlsx".replace(" ", "_")
    file_path = os.path.join(frappe.utils.get_site_path(), 'private', 'files', file_name)
    wb.save(file_path)

    # Guardar el archivo en la base de datos de Frappe
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "is_private": 1,
        "file_url": f"/private/files/{file_name}"
    })
    file_doc.save(ignore_permissions=True)

    return file_doc.file_url


def execute(filters: dict | None = None):
    """
    Punto de entrada principal para ejecutar el informe. Esta función recibe los filtros proporcionados
    por el usuario y genera el informe correspondiente.
    
    Args:
        filters (dict | None): Diccionario de filtros proporcionados por el usuario para filtrar los datos.
    
    Returns:
        tuple: Una tupla con las columnas y los datos del informe.
    """
    # Definir columnas para el informe y una lista vacía para los datos
    columns, data = get_columns(), []

    # Si no se proporcionan filtros, devolver columnas y datos vacíos
    if not filters:
        return columns, data

    # Crear un diccionario de filtros a partir de los valores proporcionados
    filters_dict = {
        "company": filters.get("company"),
        "from_date": filters.get("from_date"),
        "to_date": filters.get("to_date"),
        "iva_type": filters.get("iva_type"),  # Nuevo filtro para tipo de IVA
        "valor": filters.get("valor"),  # Nuevo filtro para valor positivo o negativo
    }

    # Determinar el tipo de factura para obtener los datos correspondientes
    if filters.get("type") == "Factura de Compra":
        data = get_invoice_data(filters_dict, "Purchase")
    else:
        data = get_invoice_data(filters_dict, "Sales")

    return columns, data

def get_invoice_data(filters_dict, invoice_type):
    """
    Obtiene los datos de las facturas en función de los filtros y el tipo de factura.
    
    Args:
        filters_dict (dict): Diccionario de filtros que contiene la empresa, fecha de inicio, fecha final, tipo de IVA y valor positivo o negativo.
        invoice_type (str): Tipo de factura, puede ser "Purchase" (Compra) o "Sales" (Venta).
    
    Returns:
        list: Lista de diccionarios con los datos de las facturas ajustados según las plantillas de impuestos.
    """
    # Definir condiciones comunes para las facturas según los filtros
    conditions = [
        "inv.company = %(company)s",
        "inv.posting_date >= %(from_date)s",
        "inv.posting_date <= %(to_date)s",
        "inv.docstatus = 1",  # Solo incluir facturas que están enviadas (docstatus = 1)
    ]

    # Configurar condiciones específicas para facturas de compra o venta
    if invoice_type == "Purchase":
        conditions.append("inv.custom_intracomunitaria = 0")  # Excluir facturas intracomunitarias
        tax_condition = "tax.account_head LIKE '472%%'"  # Condición para cuentas de impuestos
        invoice_table = "tabPurchase Invoice"
        invoice_item_table = "tabPurchase Invoice Item"
        party_table = "tabSupplier"
        party_field = "supplier"
        tax_table = "tabPurchase Taxes and Charges"
    else:
        tax_condition = "tax.account_head LIKE '477%%'"  # Condición para cuentas de impuestos
        invoice_table = "tabSales Invoice"
        invoice_item_table = "tabSales Invoice Item"
        party_table = "tabCustomer"
        party_field = "customer"
        tax_table = "tabSales Taxes and Charges"

    # Unir las condiciones para utilizarlas en la consulta SQL
    conditions_query = " AND ".join(conditions)

    # Definir la consulta SQL para obtener datos de las facturas
    query = f"""
        SELECT
            party.{party_field}_name AS supplier,
            party.tax_id AS cif,
            pii.base_net_amount AS base_net_amount,
            pii.item_tax_template AS item_tax_template,
            tax.account_head AS account_head,
            tax.rate AS tax_rate
        FROM
            `{invoice_table}` AS inv
        LEFT JOIN 
            `{invoice_item_table}` AS pii ON inv.name = pii.parent
        LEFT JOIN
            `{party_table}` AS party ON inv.{party_field} = party.name
        LEFT JOIN
            `{tax_table}` AS tax ON tax.parent = inv.name AND pii.parent = tax.parent
        WHERE
            {conditions_query}
            AND {tax_condition}
    """

    # Ejecutar la consulta SQL con los filtros proporcionados y devolver los resultados
    data = frappe.db.sql(query, filters_dict, as_dict=True)
    return adjust_for_item_tax_templates(data, filters_dict.get("iva_type"), filters_dict.get("valor"))

def adjust_for_item_tax_templates(data, iva_type_filter, valor_filter):
    """
    Ajusta los datos de las facturas en función de las plantillas de impuestos de los artículos y el filtro de valor (positivo o negativo).
    
    Args:
        data (list): Lista de diccionarios que representan los datos de las facturas obtenidas.
        iva_type_filter (str): Filtro que indica el tipo de IVA a incluir (21, 10, 4).
        valor_filter (str): Filtro que indica si se deben incluir valores "Positiva" o "Negativa".
    
    Returns:
        list: Lista de diccionarios con los datos ajustados.
    """
    adjusted_data = {}
    for row in data:
        base_amount = row.get("base_net_amount", 0)
        tax_rate = row.get("tax_rate", 0)

        # Ajustar el porcentaje de impuestos si se proporciona una plantilla de impuestos
        if row.get("item_tax_template"):
            tax_template = frappe.get_doc("Item Tax Template", row["item_tax_template"])
            for tax in tax_template.taxes:
                tax_rate = tax.tax_rate

        if abs(int(tax_rate)) == 0:
            continue
        # Asegurarse de que el tipo de IVA coincida con el filtro, incluyendo casos con IVA negativo
        if iva_type_filter is not None and abs(int(tax_rate)) != int(iva_type_filter):
            continue

        # Asegurarse de que el valor base sea positivo o negativo según el filtro de valor
        if valor_filter == "Positiva" and base_amount <= 0:
            continue
        elif valor_filter == "Negativa" and base_amount >= 0:
            continue

        # Crear una clave única basada en proveedor, CIF y porcentaje de IVA
        key = (row["supplier"], row["cif"], tax_rate)

        # Agregar datos ajustados a la estructura resultante
        if key not in adjusted_data:
            adjusted_data[key] = {
                "supplier": row["supplier"],
                "cif": row["cif"],
                "porcentaje": tax_rate,
                "total_base": 0,
                "total_iva": 0,
            }
        adjusted_data[key]["total_base"] += base_amount
        adjusted_data[key]["total_iva"] += (base_amount * tax_rate) / 100

        if valor_filter == "Negativa" and adjusted_data[key]["total_iva"] > 0:
            adjusted_data[key]["total_iva"] = -adjusted_data[key]["total_iva"]
        if valor_filter == "Negativa" and adjusted_data[key]["porcentaje"] < 0:
            adjusted_data[key]["porcentaje"] = abs(adjusted_data[key]["porcentaje"])

    # Devolver los datos ajustados como una lista de diccionarios
    return list(adjusted_data.values())


def get_columns():
    """
    Define las columnas del informe que se mostrará al usuario.
    
    Returns:
        list: Lista de diccionarios que define las columnas del informe.
    """
    return [
        {"label": _( "Proveedor/Cliente"), "fieldtype": "Link", "fieldname": "supplier", "options": "Supplier", "width": 200},
        {"label": _( "CIF"), "fieldtype": "Data", "fieldname": "cif", "width": 120},
        {"label": _( "Total Base"), "fieldtype": "Currency", "fieldname": "total_base", "width": 120},
        {"label": _( "Porcentaje de IVA"), "fieldname": "porcentaje", "fieldtype": "Data", "width": 100},
        {"label": _( "Total IVA"), "fieldtype": "Currency", "fieldname": "total_iva", "width": 120},
    ]
