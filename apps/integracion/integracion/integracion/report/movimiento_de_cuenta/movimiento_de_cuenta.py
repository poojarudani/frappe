# Movimiento de Cuenta Script Report en Frappe

import frappe
from frappe.utils import flt, getdate
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os


@frappe.whitelist()
def export_to_excel(filters=None):
    if not filters:
        return

    # Convertir los filtros a un diccionario si es un string JSON
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)

    # Obtener columnas y datos del reporte
    columns, data = execute(filters)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimiento de Cuenta"

    # Estilos
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezado del reporte
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="Movimiento de Cuenta").font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Empresa
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=f"Empresa: {filters.get('company')}").font = Font(size=12, bold=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='left')

    # Periodo
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(columns))
    ws.cell(row=3, column=1, value=f"Periodo: {filters.get('from_date')} a {filters.get('to_date')}").font = Font(size=12, bold=True)
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='left')

    # Encabezado de la tabla
    header_row = 5
    for col_num, column in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_num, value=column["label"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = gray_fill
        cell.border = thin_border

    # Agregar los datos
    for row_num, row_data in enumerate(data, header_row + 1):
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data.get(column["fieldname"]))
            cell.border = thin_border

    # Guardar el archivo Excel
    file_name = f"Movimiento_de_Cuenta_{filters.get('from_date')}_a_{filters.get('to_date')}.xlsx".replace(" ", "_")
    file_path = os.path.join(frappe.utils.get_site_path(), 'private', 'files', file_name)
    wb.save(file_path)

    # Guardar el archivo en la base de datos de Frappe
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "is_private": 1,
        "file_url": f"/private/files/{file_name}"
    })
    file_doc.save(ignore_permissions=True)

    return file_doc.file_url

def execute(filters=None):
    columns, data = get_columns(filters), []
    
    if not filters:
        return columns, data

    # Obtener los filtros
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    accounts = filters.get("account")  # Ahora es una lista de nombres de cuentas
    company = filters.get("company")
    is_group = filters.get("is_group")
    account_type = filters.get("account_type")
    totals = filters.get("totals")

    conditions = []

    # Filtros de fecha
    if from_date:
        conditions.append("gl.posting_date >= %(from_date)s")
    if to_date:
        conditions.append("gl.posting_date <= %(to_date)s")

    # Filtrar por compañía
    if company:
        conditions.append("gl.company = %(company)s")

    # Filtrar por cuentas seleccionadas
    if accounts:
        # Si es un grupo, obtenemos las cuentas hijas
        account_names = []
        for account in accounts:
            account_doc = frappe.get_doc("Account", account)  # `account` es el nombre directamente
            if account_doc.is_group:
                child_accounts = frappe.get_all("Account", filters={"parent_account": account_doc.name, "company": company}, fields=["name"])
                account_names.extend([acc["name"] for acc in child_accounts])
            else:
                account_names.append(account_doc.name)
        
        if account_names:
            conditions.append("gl.account IN %(account_names)s")
            filters["account_names"] = tuple(account_names)

    # Filtros adicionales según el tipo de cuenta
    if account_type == "Empleado":
        conditions.append("acc.custom_empleado IS NOT NULL")
    elif account_type == "Profesional":
        conditions.append("LOWER(acc.account_name) LIKE %(account_type_profesional)s")
        filters["account_type_profesional"] = '%profesional%'
    elif account_type == "Alquiler":
        conditions.append("LOWER(acc.account_name) LIKE %(account_type_alquiler)s")
        filters["account_type_alquiler"] = '%alquiler%'

    # Construir la condición para la consulta
    conditions_str = " AND ".join(conditions)

    # Consultar GL Entry con las condiciones y unir con Account para obtener el account_number
    if totals:
        query = f"""
            SELECT 
                gl.account, 
                SUM(gl.debit) AS debit, 
                SUM(gl.credit) AS credit,
                COUNT(gl.name) AS num_movements,
                acc.account_number
            FROM 
                `tabGL Entry` AS gl
            LEFT JOIN
                `tabAccount` AS acc ON gl.account = acc.name
            WHERE 
                {conditions_str}
            GROUP BY 
                gl.account, acc.account_number
            ORDER BY 
                acc.account_number ASC
        """
    else:
        query = f"""
            SELECT 
                gl.posting_date, gl.account, gl.debit, gl.credit,
                gl.voucher_type, gl.voucher_no, gl.against, gl.remarks,
                acc.account_number
            FROM 
                `tabGL Entry` AS gl
            LEFT JOIN
                `tabAccount` AS acc ON gl.account = acc.name
            WHERE 
                {conditions_str}
            ORDER BY 
                acc.account_number ASC, gl.posting_date ASC
        """

    # Ejecutar la consulta
    data = frappe.db.sql(query, filters or {}, as_dict=True)

    # Devolver las columnas y los datos
    return columns, data

def get_columns(filters=None):
    return [
        {"label": "Fecha de Publicación", "fieldname": "posting_date", "fieldtype": "Date", "width": 100},
        {"label": "Cuenta", "fieldname": "account", "fieldtype": "Link", "options": "Account", "width": 150},
        {"label": "Débito", "fieldname": "debit", "fieldtype": "Currency", "width": 120},
        {"label": "Crédito", "fieldname": "credit", "fieldtype": "Currency", "width": 120},
        {"label": "Tipo de Comprobante", "fieldname": "voucher_type", "fieldtype": "Data", "width": 120},
        {"label": "Número de Comprobante", "fieldname": "voucher_no", "fieldtype": "Data", "width": 120},
        {"label": "Contra Cuenta", "fieldname": "against", "fieldtype": "Data", "width": 120},
        {"label": "Observaciones", "fieldname": "remarks", "fieldtype": "Data", "width": 200},
        {"label": "Número de Cuenta", "fieldname": "account_number", "fieldtype": "Data", "width": 120}
    ] if not filters.get("totals") else [
        {"label": "Cuenta", "fieldname": "account", "fieldtype": "Link", "options": "Account", "width": 150},
        {"label": "Débito", "fieldname": "debit", "fieldtype": "Currency", "width": 120},
        {"label": "Crédito", "fieldname": "credit", "fieldtype": "Currency", "width": 120},
        {"label": "Número de Movimientos", "fieldname": "num_movements", "fieldtype": "Int", "width": 120},
        {"label": "Número de Cuenta", "fieldname": "account_number", "fieldtype": "Data", "width": 120}
    ]
